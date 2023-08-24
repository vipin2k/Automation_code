import os

import pytest
from selenium import webdriver
import openpyxl
from pathlib import Path
import time
import string
import random
from selenium.webdriver.common.by import By
from pageObjects.Discharge_Request import DischargePage
from pageObjects.Medical_IP_Request_CreationPage import Medical_IP_Request_CreationPage
from pageObjects.Medical_OP_Request_CreationPage import Medical_OP_Request_CreationPage
from pageObjects.MemberSearchPage import MemberSearchPage
from pageObjects.MemberPage import MemberPage
from pageObjects.EnrollmentPage import EnrollmentPage
from pageObjects.RequestHomePage import RequestPage
from pageObjects.Task_CreationPage import TaskPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen


class Test_002_Creating_Member:
    baseurl = ReadConfig.getApplicationURl()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    firstname = ReadConfig.getFirstname()
    lastname = ReadConfig.getLastname()
    member_id = ReadConfig.getMember_id()
    DOB = ReadConfig.getMember_DOB()
    Effective_date = ReadConfig.getEffective_date()
    search_id = ReadConfig.getSearchMemberid()
    Admit_date = ReadConfig.getAdmit_date()
    discharge_date = ReadConfig.getdischarge_date()
    unit_date = ReadConfig.getunit_date()
    facility = ReadConfig.getfacility()
    start_date = ReadConfig.getstart_date()
    To_date = ReadConfig.getTo_date()
    logger = LogGen.loggen()
    # @pytest.mark.sanity
    def test_createMember(self, setup):
        self.logger.info("**************Verifying Home Page********************")
        self.driver = setup
        self.driver.get(self.baseurl)
        self.driver.maximize_window()
        self.mp = MemberPage(self.driver)
        self.mp.setUserName(self.username)
        self.mp.setPassword(self.password)
        self.mp.clickLogin()
        self.mp.clickSearch()
        self.mp.clickMember()
        self.mp.clickNewMember()
        self.mp.setFirstName(self.firstname)
        self.mp.setLastName(self.lastname)
        self.Number = 5
        value = ''.join(random.choices(string.ascii_uppercase + string.digits, k=self.Number))
        print("Member patient ID : " + str(value))
        self.mp.setMemberId(self.member_id + str(value))
        self.mp.setDob(self.DOB)
        time.sleep(5)
        self.mp.clickGender()
        self.ep = EnrollmentPage(self.driver)
        self.logger.info("**************Verifying Enrollment Page********************")
        self.ep.clickEnrollment()
        self.ep.clickAdd()
        self.ep.clickThisMember()
        self.ep.clickRelationship()
        time.sleep(1)
        self.ep.clickRelationship_type()
        self.ep.setEffective_date(self.Effective_date)
        self.ep.clickSave()
        self.msg = MemberSearchPage(self.driver)
        time.sleep(2)
        self.msg.clickSearchIcon()
        self.msg.clickMemberIcon()
        self.msg.setSearchMemberid(self.search_id + str(value))
        self.msg.clickSearchAerial()
        time.sleep(3)
        search_id = value
        sample_xpath = f'//a[@id="memberSearchResultsForm:memberSearchResultsTable:searchResultsTable:0:memberDemographics" and text()="{value}"]'
        self.txt = self.driver.find_element(By.XPATH, sample_xpath).text
        # print(self.txt)
        if search_id:
            assert search_id == self.txt
            self.logger.info("*******ACM user Login Successful- Test Passed*******")
        else:
            self.logger.error("*******ACM user Login Successful- Test Failed*******")
        self.rp = RequestPage(self.driver)
        self.rp.clickCmv()
        time.sleep(4)
        self.rp.clickProgram()
        time.sleep(2)
        self.mrop = Medical_OP_Request_CreationPage(self.driver)
        self.logger.info("**************Verifying Request Creation Page********************")
        time.sleep(2)
        self.mrop.click_Medical_OP()
        time.sleep(2)
        self.mrop.click_Treatment_setting()
        time.sleep(1)
        self.mrop.click_Treatment_setting_type()
        time.sleep(1)
        self.mrop.clickurgency()
        time.sleep(1)
        self.mrop.click_emergency()
        time.sleep(1)
        self.mrop.clickproviders_type()
        time.sleep(2)
        self.mrop.click_facility()
        time.sleep(2)
        self.mrop.setfacility(self.facility)
        time.sleep(2)
        self.mrop.clicksearch_aerial()
        time.sleep(1)
        self.mrop.clickadd_facility()
        time.sleep(2)
        self.mrop.click_select_facility()
        time.sleep(2)
        self.mrop.click_select_diagnoses()
        time.sleep(2)
        self.mrop.clickdiagnoses_dropdown()
        time.sleep(2)
        self.mrop.click_select_diagnoses_dropdown()
        time.sleep(2)
        self.mrop.click_select_diagnosis_add()
        time.sleep(3)
        self.mrop.clickselect_services()
        time.sleep(2)
        self.mrop.clickservice_dropdown()
        time.sleep(1)
        self.mrop.click_select_service_dropdown()
        time.sleep(3)
        self.mrop.click_select_service_add()
        time.sleep(4)
        self.mrop.click_select_service_scheme()
        time.sleep(3)
        self.mrop.click_service_treatment_dropdown()
        time.sleep(2)
        self.mrop.click_service_treatment_type_dropdown_value()
        time.sleep(2)
        self.mrop.set_service_start_date(self.start_date)
        time.sleep(1)
        self.mrop.clickrequest_units()
        time.sleep(2)
        self.mrop.setUnits(self.unit_date)
        time.sleep(3)
        self.mrop.clickreview_tab()
        self.mrop.clickreview_tab()
        time.sleep(8)
        self.mrop.clickselect_mcg_page()
        time.sleep(4)
        self.mrop.click_mcg_episode_page()
        time.sleep(5)
        self.mrop.set_review_TODATE(self.To_date)
        time.sleep(1)
        self.mrop.clickreview_status_dropdown()
        self.mrop.clickreview_status_dropdown()
        time.sleep(2)
        self.mrop.click_status_dropdown_value()
        time.sleep(2)
        self.mrop.click_Add_outcome()
        time.sleep(5)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_Medical_One_Open_task_One_close_task/Request.png")
        self.mrop.clicksave()
        self.mrop.clicktitle_home()
        self.msp = MemberSearchPage(self.driver)
        self.logger.info("**************Verifying Member Search Page********************")
        self.msp.clickSearchIcon()
        self.msp.clickMemberIcon()
        self.msp.clearMemberid()
        self.msp.setSearchMemberid(self.search_id + str(value))
        self.msp.clickSearchAerial()
        time.sleep(2)
        self.msp.clickCmv()
        time.sleep(2)
        # self.Admit_date = "03/22/2023"
        # admit = self.Admit_date
        # self.date = self.driver.find_element(By.XPATH, "(//*[@class='inpatient addDynamicTooltip'])[1]").text
        self.rp.clickcreated_request()
        time.sleep(2)
        request_id = self.driver.find_element(By.XPATH, '(//*[@class="ui-panelgrid-cell"])[30]').text
        # print(request_id)
        my_file = Path(
            "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        if (request_id != None):
            if my_file.exists():
                bk = openpyxl.load_workbook(
                    "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
                s = bk.active
                s.cell(row=8, column=5).value = request_id.split(" ")[0]
                s.cell(row=8, column=6).value = "Pass"
                bk.save(
                    "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            else:
                print("Excel file not found")
        else:
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=8, column=5).value = "None"
            s.cell(row=8, column=6).value = "Fail"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")

        print(request_id.split(" ")[0])
        self.mrop.clickextend()
        time.sleep(10)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_Medical_One_Open_task_One_close_task/Extension.png")
        self.mrop.clickrequest_units()
        time.sleep(2)
        self.mrop.setUnits(self.unit_date)
        time.sleep(2)
        self.mrop.clicksave()
        self.mrop.clicksave()
        time.sleep(3)
        self.mrop.clicktitle_home()
        self.msp = MemberSearchPage(self.driver)
        self.logger.info("**************Verifying Member Search Page********************")
        self.msp.clickSearchIcon()
        self.msp.clickMemberIcon()
        self.msp.clearMemberid()
        self.msp.setSearchMemberid(self.search_id + str(value))
        self.msp.clickSearchAerial()
        time.sleep(2)
        self.msp.clickCmv()
        time.sleep(2)
        self.rp = RequestPage(self.driver)
        time.sleep(2)
        # request_id = self.driver.find_element(By.XPATH, '(//*[@class="ui-panelgrid-cell"])[30]').text
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clickpriority()
        time.sleep(2)
        self.tp.clickpriority_type()
        time.sleep(3)
        self.tp.clicksave_exit()
        time.sleep(1)
        self.driver.refresh()
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clickpriority()
        time.sleep(2)
        self.tp.clickpriority_type()
        time.sleep(10)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_Medical_One_Open_task_One_close_task/Task.png")
        self.tp.clicksave_exit()
        time.sleep(1)
        self.driver.refresh()
        self.rp = RequestPage(self.driver)
        self.rp.clickcreated_request()
        time.sleep(2)
        request_id = self.driver.find_element(By.XPATH, '(//*[@class="ui-panelgrid-cell"])[30]').text
        print(request_id.split(" ")[0])
        Extension_verification = self.driver.find_element(By.XPATH,
                                                          "//*[@id='cmvCenterForm:cmvItemsTabView:0:requestSummaryTabView:summaryServicesDataListOPP:0:subServicesAccordionPanel:1:requestDetailEditExtensionCommandButton']/span").text
        print(Extension_verification)

        if (Extension_verification == "EDIT EXTENSION"):
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=8, column=7).value = request_id.split(" ")[0]
            s.cell(row=8, column=8).value = "Pass"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        else:
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=8, column=7).value = "None"
            s.cell(row=8, column=8).value = "Fail"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        self.Dp = DischargePage(self.driver)
        time.sleep(2)
        self.Dp.click_close_task()
        time.sleep(10)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_Medical_One_Open_task_One_close_task/Discharge.png")
        # self.Dp.click_close_task_dropdown()
        # time.sleep(2)
        # self.Dp.click_dropdown_close()
        time.sleep(2)
        self.Dp.click_close_wizard()
        time.sleep(5)
        # self.Dp.click_discharge_button()
        # time.sleep(2)
        # self.Dp.click_discharge_date()
        # time.sleep(1)
        # self.Dp.set_discharge_date(self.discharge_date)
        # time.sleep(2)
        # self.Dp.click_disposition_button()
        # self.Dp.click_disposition_button()
        # time.sleep(2)
        # self.Dp.click_disposition_type()
        # time.sleep(2)
        # self.Dp.click_discharge()
        # time.sleep(4)
        # self.Dp.click_discharge_Task()
        # time.sleep(2)
        # task_closed = self.driver.find_element(By.XPATH,
        #                                "//*[@id='cmvCenterForm:cmvItemsTabView:1:filteredTaskAccordianPanel:closedTaskListTableFiltered_data']/tr[1]/td[3]").text
        # request_closed = self.driver.find_element(By.XPATH, "//*[@id='cmvCenterForm:cmvItemsTabView:1:filteredTaskAccordianPanel:closedTaskListTableFiltered_data']/tr[2]/td[3]").text
        # try:
        #     assert task_closed == request_closed
        #     self.logger.info("*******ACM task and request Discharge successfully - Test Passed*******")
        #     print("Request and Task discharge text matched :" + task_closed)
        # except:
        #     self.logger.error("******ACM task and request Discharge fail- Test Failed*******")


