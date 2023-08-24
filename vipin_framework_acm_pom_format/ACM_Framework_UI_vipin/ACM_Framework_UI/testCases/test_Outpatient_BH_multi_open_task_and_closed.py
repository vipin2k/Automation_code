import os

import pytest
from selenium import webdriver
import time
import string
import random
import openpyxl
from pathlib import Path
from selenium.webdriver.common.by import By
from pageObjects.Discharge_Request import DischargePage
from pageObjects.Medical_OP_Request_CreationPage import Medical_OP_Request_CreationPage
from pageObjects.MemberSearchPage import MemberSearchPage
from pageObjects.MemberPage import MemberPage
from pageObjects.EnrollmentPage import EnrollmentPage
from pageObjects.RequestHomePage import RequestPage
from pageObjects.Task_CreationPage import TaskPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from pageObjects.BehavioralHealth_Prior_Auth_Request_CreationPage import BehavioralHealth_Prior_Auth
from pageObjects.Medical_OPBH_Request_CreationPage import Medical_OPBH_Request_CreationPage

class Test_004Creating_Member:
    baseurl = ReadConfig.getApplicationURl()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    firstname = ReadConfig.getFirstname()
    lastname = ReadConfig.getLastname()
    member_id = ReadConfig.getMember_id()
    DOB = ReadConfig.getMember_DOB()
    Effective_date = ReadConfig.getEffective_date()
    search_id = ReadConfig.getSearchMemberid()
    Admit_date = ReadConfig.getAdmit_date()
    discharge_date = ReadConfig.getdischarge_date()
    existing_member_id = ReadConfig.getexisting_member_id()
    unit_date = ReadConfig.getunit_date()
    facility = ReadConfig.getfacility()
    startdate = ReadConfig.getstart_date()
    requested_units = ReadConfig.getrequestedunits()
    Description = ReadConfig.getDescription()
    To_Date = ReadConfig.getTo_date()
    logger = LogGen.loggen()


    def test_createMember(self, setup):
        self.logger.info("**************Creating_Member********************")
        self.logger.info("**************Verifying Home Page********************")
        self.driver = setup
        self.driver.get(self.baseurl)
        self.driver.maximize_window()
        self.mp = MemberPage(self.driver)
        self.mp.setUserName(self.username)
        self.mp.setPassword(self.password)
        self.mp.clickLogin()
        self.mp.clickSearch()
        self.mp.clickMember()
        self.mp.clickNewMember()
        self.mp.setFirstName(self.firstname)
        self.mp.setLastName(self.lastname)
        self.Number = 5
        value = ''.join(random.choices(string.ascii_uppercase + string.digits, k=self.Number))
        print("Member patient ID : " + str(value))
        self.mp.setMemberId(self.member_id + str(value))
        self.mp.setDob(self.DOB)
        time.sleep(3)
        self.mp.clickGender()
        self.ep = EnrollmentPage(self.driver)
        self.logger.info("**************Creating_Member********************")
        self.logger.info("**************Verifying Enrollment Page********************")
        self.ep.clickEnrollment()
        self.ep.clickAdd()
        self.ep.clickThisMember()
        self.ep.clickRelationship()
        time.sleep(2)
        self.ep.clickRelationship_type()
        self.ep.setEffective_date(self.Effective_date)
        self.ep.clickSave()
        self.msg = MemberSearchPage(self.driver)
        time.sleep(2)
        self.msg.clickSearchIcon()
        self.msg.clickMemberIcon()
        self.msg.setSearchMemberid(self.search_id + str(value))
        self.msg.clickSearchAerial()
        time.sleep(3)
        search_id = value
        sample_xpath = f'//a[@id="memberSearchResultsForm:memberSearchResultsTable:searchResultsTable:0:memberDemographics" and text()="{value}"]'
        self.txt = self.driver.find_element(By.XPATH, sample_xpath).text
        if search_id:
            assert search_id == self.txt
            self.logger.info("*******ACM user Login Successful- Test Passed*******")
        else:
            self.logger.error("*******ACM user Login Successful- Test Failed*******")
        self.rp = RequestPage(self.driver)
        self.rp.clickCmv()
        time.sleep(4)
        self.rp.clickProgram()
        time.sleep(2)
        # self.mrc = Medical_IP_Request_CreationPage(self.driver)
        self.mrop = Medical_OPBH_Request_CreationPage(self.driver)
        self.logger.info("**************Creating_Member********************")
        self.logger.info("**************Verifying Request Creation Page********************")
        time.sleep(2)
        self.mrop.click_Medical_OPBH()
        time.sleep(3)
        self.mrop.click_Treatment_setting()
        time.sleep(1)
        self.mrop.click_Treatment_setting_type()
        time.sleep(1)
        self.mrop.clickurgency()
        time.sleep(1)
        self.mrop.click_emergency()
        time.sleep(1)
        self.mrop.clickproviders_type()
        time.sleep(1)
        self.mrop.click_facility()
        time.sleep(1)
        self.mrop.setfacility(self.facility)
        time.sleep(1)
        self.mrop.clicksearch_aerial()
        time.sleep(1)
        self.mrop.clickadd_facility()
        time.sleep(1)
        self.mrop.click_select_facility()
        time.sleep(2)
        self.mrop.click_Diagnoses()
        time.sleep(3)
        self.mrop.click_Diagnoses_Description(self.Description)
        time.sleep(3)
        self.mrop.click_search_diagnoses()
        time.sleep(1)
        self.mrop.clickdiagnoses_add()
        time.sleep(3)
        self.mrop.clickselect_services()
        time.sleep(3)
        self.mrop.clickservice_dropdown()
        time.sleep(3)
        self.mrop.click_select_service_dropdown()
        time.sleep(1)
        self.mrop.click_select_service_add()
        time.sleep(2)

        self.mrop.click_service_tab()
        time.sleep(2)
        self.mrop.click_Treatment_setting_service()
        time.sleep(2)
        self.mrop.click_Treatment_setting_dropdown()
        time.sleep(2)
        self.mrop.click_startdate(self.startdate)
        time.sleep(3)
        self.mrop.click_request_units_service()
        time.sleep(2)
        self.mrop.set_request_units_service(self.requested_units)
        time.sleep(2)

        self.mrop.clickReview_tab()
        self.mrop.clickReview_tab()
        time.sleep(8)
        self.mrop.clickmcg_page()
        time.sleep(4)
        self.mrop.clickmcg_episode_page()
        time.sleep(5)
        self.mrop.set_review_To_Date(self.To_Date)
        time.sleep(2)
        self.mrop.clickreview_status_dropdown()
        self.mrop.clickreview_status_dropdown()
        time.sleep(2)
        self.mrop.click_dropdown_value()
        time.sleep(5)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_BH_multi_open_closed_task/Request.png")
        self.mrop.clickAdd_outcome()
        time.sleep(2)
        self.mrop.clicksave()
        time.sleep(5)
        # print(os.getcwd())
        time.sleep(3)
        request_id = self.driver.find_element(By.XPATH, '(//*[@class="ui-panelgrid-cell"])[30]').text
        # print(request_id)
        row_number = None
        my_file = Path(
            "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        if (request_id != None):
            if my_file.exists():
                bk = openpyxl.load_workbook(
                    "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
                s = bk.active
                s.cell(row=9, column=5).value = request_id.split(" ")[0]
                s.cell(row=9, column=6).value = "Pass"
                bk.save(
                    "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            else:
                print("Excel file not found")
        else:
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=9, column=5).value = "None"
            s.cell(row=9, column=6).value = "Fail"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")

        print(request_id.split(" ")[0])
        time.sleep(2)
        self.mrop.clickextend()
        time.sleep(1)
        self.mrop.click_request_units_service()
        time.sleep(1)
        self.mrop.set_request_units(self.requested_units)
        time.sleep(2)
        # print(os.getcwd())
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_BH_multi_open_closed_task/Extension.png")
        time.sleep(3)
        # Extension
        self.mrop.clicksave()
        self.mrop.clicksave()
        time.sleep(5)
        self.mrop.clicktitle_home()
        self.msp = MemberSearchPage(self.driver)
        self.logger.info("**************Creating_Member********************")
        self.logger.info("**************Verifying Member Search Page********************")
        self.msp.clickSearchIcon()
        self.msp.clickMemberIcon()
        self.msp.clearMemberid()
        self.msp.setSearchMemberid(self.search_id + str(value))
        self.msp.clickSearchAerial()
        time.sleep(2)
        self.msp.clickCmv()
        time.sleep(2)
        self.rp.clickcreated_request()
        time.sleep(2)

        self.rp = RequestPage(self.driver)
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(5)
        # print(os.getcwd())
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_BH_multi_open_closed_task/Task_1.png")
        time.sleep(3)
        # self.tp.clickpriority()
        # time.sleep(2)
        # self.tp.clickpriority_type()
        # time.sleep(3)
        self.tp.clicksave_exit()
        time.sleep(3)
        self.driver.refresh()
        time.sleep(5)
        self.rp = RequestPage(self.driver)
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(5)
        # print(os.getcwd())
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_BH_multi_open_closed_task/Task_2.png")
        time.sleep(3)
        # self.tp.clickpriority()
        # time.sleep(2)
        # self.tp.clickpriority_type()
        # time.sleep(3)
        self.tp.clicksave_exit()
        time.sleep(3)
        self.driver.refresh()
        time.sleep(5)
        self.rp = RequestPage(self.driver)
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clicksave_exit()
        time.sleep(1)
        self.driver.refresh()
        time.sleep(5)
        self.rp = RequestPage(self.driver)
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(3)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clicksave_exit()
        time.sleep(3)
        self.driver.refresh()
        time.sleep(5)
        self.rp = RequestPage(self.driver)
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clicksave_exit()
        time.sleep(3)
        self.driver.refresh()
        time.sleep(3)
        self.rp.clickcreated_request()
        time.sleep(2)
        Extension_verification = self.driver.find_element(By.XPATH,
                                                          "//*[@id='cmvCenterForm:cmvItemsTabView:0:requestSummaryTabView:summaryServicesDataListOPP:0:subServicesAccordionPanel:1:requestDetailEditExtensionCommandButton']/span").text
        print(Extension_verification)
        if (Extension_verification == "EDIT EXTENSION"):
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=9, column=7).value = request_id.split(" ")[0]
            s.cell(row=9, column=8).value = "Pass"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        else:
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=9, column=7).value = "None"
            s.cell(row=9, column=8).value = "Fail"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        time.sleep(2)
        # closetask
        self.tp.close_task_1()
        time.sleep(3)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_BH_multi_open_closed_task/Discharge_1.png")
        time.sleep(3)
        self.tp.close_task_wizard_button()
        time.sleep(2)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Outpatient_BH_multi_open_closed_task/Discharge_2.png")
        time.sleep(3)
        self.tp.close_task_2()
        time.sleep(2)
        self.tp.close_task_wizard_button()
        time.sleep(3)

        # print(os.getcwd())
