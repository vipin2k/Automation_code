import os

import pytest
from selenium import webdriver
import openpyxl
from pathlib import Path
import time
import string
import random
from selenium.webdriver.common.by import By
from pageObjects.Discharge_Request import DischargePage
from pageObjects.Medical_IP_Request_CreationPage import Medical_IP_Request_CreationPage
from pageObjects.MemberSearchPage import MemberSearchPage
from pageObjects.MemberPage import MemberPage
from pageObjects.EnrollmentPage import EnrollmentPage
from pageObjects.Referral_Medical_Request_CreationPage import Referral_Medical_Request_CreationPage
from pageObjects.RequestHomePage import RequestPage
from pageObjects.Task_CreationPage import TaskPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen


class Test_002_Creating_Member:
    baseurl = ReadConfig.getApplicationURl()
    username = ReadConfig.getUsername()
    password = ReadConfig.getPassword()
    firstname = ReadConfig.getFirstname()
    lastname = ReadConfig.getLastname()
    member_id = ReadConfig.getMember_id()
    DOB = ReadConfig.getMember_DOB()
    Effective_date = ReadConfig.getEffective_date()
    search_id = ReadConfig.getSearchMemberid()
    Admit_date = ReadConfig.getAdmit_date()
    discharge_date = ReadConfig.getdischarge_date()
    unit_date = ReadConfig.getunit_date()
    start_date = ReadConfig.getstart_date()
    To_date = ReadConfig.getTo_date()
    requested_units = ReadConfig.getrequestedunits()
    facility = ReadConfig.getfacility()
    logger = LogGen.loggen()
    # @pytest.mark.sanity
    def test_createMember(self, setup):
        self.logger.info("**************Verifying Home Page********************")
        self.driver = setup
        self.driver.get(self.baseurl)
        self.driver.maximize_window()
        self.mp = MemberPage(self.driver)
        self.mp.setUserName(self.username)
        self.mp.setPassword(self.password)
        self.mp.clickLogin()
        self.mp.clickSearch()
        self.mp.clickMember()
        self.mp.clickNewMember()
        self.mp.setFirstName(self.firstname)
        self.mp.setLastName(self.lastname)
        self.Number = 5
        value = ''.join(random.choices(string.ascii_uppercase + string.digits, k=self.Number))
        print("Member Patient ID : " + str(value))
        self.mp.setMemberId(self.member_id + str(value))
        self.mp.setDob(self.DOB)
        time.sleep(5)
        self.mp.clickGender()
        self.ep = EnrollmentPage(self.driver)
        self.logger.info("**************Verifying Enrollment Page********************")

    # def test_Enrollment(self, setup):
        self.ep.clickEnrollment()
        self.ep.clickAdd()
        self.ep.clickThisMember()
        self.ep.clickRelationship()
        time.sleep(1)
        self.ep.clickRelationship_type()
        self.ep.setEffective_date(self.Effective_date)
        self.ep.clickSave()
        self.msg = MemberSearchPage(self.driver)
        time.sleep(2)
        self.msg.clickSearchIcon()
        self.msg.clickMemberIcon()
        self.msg.setSearchMemberid(self.search_id + str(value))
        self.msg.clickSearchAerial()
        time.sleep(3)
        search_id = value
        sample_xpath = f'//a[@id="memberSearchResultsForm:memberSearchResultsTable:searchResultsTable:0:memberDemographics" and text()="{value}"]'
        self.txt = self.driver.find_element(By.XPATH, sample_xpath).text
        print(self.txt)
        if search_id:
            assert search_id == self.txt
            self.logger.info("*******ACM user Login Successful- Test Passed*******")
        else:
            self.logger.error("*******ACM user Login Successful- Test Failed*******")
        self.rp = RequestPage(self.driver)
        self.rp.clickCmv()
        time.sleep(4)
        self.rp.clickProgram()
        time.sleep(2)
        self.RMR = Referral_Medical_Request_CreationPage(self.driver)
        self.logger.info("**************Verifying Request Creation Page********************")
        time.sleep(2)
        self.RMR.clickMedical()
        time.sleep(2)
        self.RMR.clickTreatment()
        time.sleep(2)
        self.RMR.clickTreatmentsetting()
        time.sleep(2)
        self.RMR.clickTreatment_type()
        self.RMR.clickTreatment_type()
        time.sleep(2)
        self.RMR.clickRequest_treatment_type()
        time.sleep(2)
        self.RMR.click_request_units()
        time.sleep(2)
        self.RMR.set_requested_units(self.requested_units)
        time.sleep(1)
        self.RMR.click_start_date()
        time.sleep(2)
        self.RMR.set_start_date(self.start_date)
        time.sleep(2)
        self.RMR.clickproviders_type()
        self.RMR.clickproviders_type()
        time.sleep(2)
        self.RMR.click_facility()
        self.RMR.setfacility(self.facility)
        self.RMR.clicksearch_aerial()
        time.sleep(2)
        self.RMR.clickadd_facility()
        time.sleep(3)
        self.RMR.click_select_referring_provider_facility()
        time.sleep(2)
        self.RMR.click_select_servicing_provider_facility()
        time.sleep(1)
        self.RMR.click_select_diagnoses()
        time.sleep(1)
        self.RMR.clickdiagnoses_dropdown()
        time.sleep(2)
        self.RMR.click_select_diagnoses_dropdown()
        time.sleep(2)
        self.RMR.click_select_diagnosis_add()
        time.sleep(3)
        self.RMR.clickselect_services()
        time.sleep(2)
        self.RMR.clickservice_dropdown()
        time.sleep(1)
        self.RMR.click_select_service_dropdown()
        time.sleep(3)
        self.RMR.click_select_service_add()
        # time.sleep(4)
        # self.RMR.click_select_service_scheme()
        # self.RMR.click_select_service_scheme()
        # time.sleep(3)
        # self.RMR.click_servicing_provider_dropdown()
        # time.sleep(2)
        # self.RMR.click_servicing_provider_dropdown_value()
        time.sleep(2)
        self.RMR.clickReview_tab()
        # self.RMR.clickReview_tab()
        time.sleep(2)
        self.RMR.click_To_date()
        time.sleep(2)
        self.RMR.set_To_date(self.To_date)
        time.sleep(2)
        self.RMR.clickreview_status_dropdown()
        self.RMR.clickreview_status_dropdown()
        time.sleep(2)
        self.RMR.click_dropdown_value()
        time.sleep(2)
        self.RMR.clickAdd_outcome()
        time.sleep(5)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Referral_medical_only_closed/Request.png")
        self.RMR.clicksave()
        self.RMR.clicktitle_home()
        self.msp = MemberSearchPage(self.driver)
        self.logger.info("**************Verifying Member Search Page********************")
        self.msp.clickSearchIcon()
        self.msp.clickMemberIcon()
        self.msp.clearMemberid()
        self.msp.setSearchMemberid(self.search_id + str(value))
        self.msp.clickSearchAerial()
        time.sleep(2)
        self.msp.clickCmv()
        time.sleep(2)
        # self.Admit_date = "03/22/2023"
        # admit = self.Admit_date
        # self.date = self.driver.find_element(By.XPATH, "(//*[@class='inpatient addDynamicTooltip'])[1]").text
        # time.sleep(2)
        self.rp.clickcreated_request()
        time.sleep(3)
        # date = self.date.split(" ")[0]
        # try:
        #     assert admit == date
        #     self.logger.info("*******ACM request creation date- Test Passed*******")
        #     print("current date :" + admit)
        # except:
        #     self.logger.error("******ACM request creation date- Test Failed*******")
        # time.sleep(5)
        request_id = self.driver.find_element(By.XPATH, '(//*[@class="ui-panelgrid-cell"])[30]').text
        # print(request_id)
        my_file = Path(
            "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        if (request_id != None):
            if my_file.exists():
                bk = openpyxl.load_workbook(
                    "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
                s = bk.active
                s.cell(row=10, column=5).value = request_id.split(" ")[0]
                s.cell(row=10, column=6).value = "Pass"
                bk.save(
                    "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            else:
                print("Excel file not found")
        else:
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=10, column=5).value = "None"
            s.cell(row=10, column=6).value = "Fail"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")

        print(request_id.split(" ")[0])
        self.RMR.clickextend()
        time.sleep(10)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Referral_medical_only_closed/Extension.png")
        self.RMR.click_request_units()
        time.sleep(2)
        self.RMR.set_requested_units(self.requested_units)
        time.sleep(2)
        self.RMR.clicksave()
        self.RMR.clicksave()
        time.sleep(5)
        self.RMR.clicktitle_home()
        self.msp = MemberSearchPage(self.driver)
        self.logger.info("**************Verifying Member Search Page********************")
        self.msp.clickSearchIcon()
        self.msp.clickMemberIcon()
        self.msp.clearMemberid()
        self.msp.setSearchMemberid(self.search_id + str(value))
        self.msp.clickSearchAerial()
        time.sleep(2)
        self.msp.clickCmv()
        time.sleep(2)
        self.rp = RequestPage(self.driver)
        time.sleep(2)
        self.rp.clickcreated_request()
        time.sleep(5)
        Extension_verification = self.driver.find_element(By.XPATH,
                                                          "//*[@class='ui-button-text ui-c' and text() = 'Edit Extension ']").text
        print(Extension_verification)
        if (Extension_verification == "EDIT EXTENSION"):
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=10, column=7).value = request_id.split(" ")[0]
            s.cell(row=10, column=8).value = "Pass"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
        else:
            bk = openpyxl.load_workbook(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")
            s = bk.active
            s.cell(row=10, column=7).value = "None"
            s.cell(row=10, column=8).value = "Fail"
            bk.save(
                "D:\\Users\\DNetaji\\.jenkins\\jobs\\ACM_UI-Auto_close_when_request_is_discharge\\workspace\\ACM_Framework_UI\\testCases\\Auto_Close_Scenarios_for_23.1_December_Release.xlsx")

        # request_id = self.driver.find_element(By.XPATH, '(//*[@class="ui-panelgrid-cell"])[30]').text
        time.sleep(2)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clickpriority()
        time.sleep(2)
        self.tp.clickpriority_type()
        time.sleep(3)
        self.tp.clicksave_exit()
        time.sleep(1)
        self.driver.refresh()
        time.sleep(5)
        self.rp.clicknew_item()
        time.sleep(2)
        self.rp.clicktask()
        time.sleep(2)
        self.tp = TaskPage(self.driver)
        time.sleep(3)
        self.tp.clicktask_id()
        time.sleep(1)
        self.tp.clicktask_type()
        time.sleep(2)
        self.tp.clicktask_reason()
        time.sleep(2)
        self.tp.clickreason()
        time.sleep(2)
        self.tp.clickpriority()
        time.sleep(2)
        self.tp.clickpriority_type()
        time.sleep(10)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Referral_medical_only_closed/Task.png")
        self.tp.clicksave_exit()
        time.sleep(1)
        self.driver.refresh()
        time.sleep(5)
        self.rp = RequestPage(self.driver)
        self.rp.clickcreated_request()
        time.sleep(2)

        self.Dp = DischargePage(self.driver)
        time.sleep(2)
        self.Dp.click_discharge_button()
        time.sleep(2)
        self.driver.get_screenshot_as_file(
            f"{str(os.getcwd())}/ACM_Framework_UI/Screenshots/Referral_medical_only_closed/Discharge.png")
        self.Dp.click_discharge()
        time.sleep(2)
